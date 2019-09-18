import boto3

import openpyxl
from openpyxl import load_workbook
import pandas as pd
from pandas import ExcelWriter
import io

# global variable
bucket = ''
aws_access_key_id = ''
aws_secret_access_key = ''

s3 = boto3.resource('s3')
conn = boto3.client('s3', aws_access_key_id='aws_access_key_id', aws_secret_access_key='aws_secret_access_key')
response = conn.list_buckets()


def syncExcelToS3():
    # python3 is using BytesIO
    buffer = io.BytesIO()

    df_1 = pd.DataFrame()
    df_2 = pd.DataFrame()
    df_3 = pd.DataFrame()

    df_1.to_excel(buffer, sheet_name= 'Sheet1' , index=False)
    append_df_to_excel(buffer,  df_2, sheet_name='Sheet1', startrow= 10, index=False)
    append_df_to_excel(buffer,  df_3, sheet_name='Sheet1', startrow= 20, index=False)

    session = boto3.Session(aws_access_key_id= aws_access_key_id, aws_secret_access_key= aws_secret_access_key)
    s3 = session.resource('s3')
    s3.Object(bucket, 'Full/Path').put(Body= buffer.getvalue())



def syncJsonToS3():
    jsonBuffer = io.StringIO()

    Cars = {'Brand': ['Honda Civic','Toyota Corolla','Ford Focus','Audi A4'],
            'Price': [22000,25000,27000,35000]
            }
    df = pd.DataFrame(Cars,columns= ['Brand', 'Price'])
    session = boto3.Session(aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)
    s3 = session.resource('s3')
    df.to_json(outputBuffer, orient='values');
    s3.Object(Bucket, 'PATH/TO/test.json').put(Body= jsonBuffer.getvalue())

    
    
'''
HELP Function
Append a DataFrame [df] to existing Excel file [filename]
stackoverflow: https://stackoverflow.com/questions/47737220/append-dataframe-to-excel-with-pandas
'''
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

if __name__ == '__main__':
    # append multiple dfs to different spreadsheets in one excel, then upload it to s3 using BytesIO
    syncExcelToS3();
    # create df, then upload it to s3 as a json file using StringIO
    syncJsonToS3();
